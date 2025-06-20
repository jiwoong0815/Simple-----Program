import datetime
import os
import platform
import subprocess
from decimal import Decimal
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from models import Company, InvoiceLine, Item, PriceTier # Item and PriceTier might be needed for testing or context

def create_invoice_excel(
    company: Company,
    invoice_lines: List[InvoiceLine],
    invoice_date: Optional[datetime.date] = None, # 명세서 발행일 (선택)
    # 추가적인 회사 정보 (공급자 정보 등)는 필요시 인자로 추가 가능
    # supplier_name: str = "Y준 메디컬",
    # supplier_details: Dict[str, str] = None, # 예: {"사업자번호": "123-45-67890", ...}
) -> Optional[str]:
    """
    주어진 회사 정보와 품목들로 거래명세서 Excel 파일을 생성합니다.
    레이아웃은 지정된 이미지와 유사하게 구성됩니다.

    Args:
        company: 공급받는자 회사 객체.
        invoice_lines: 명세서에 포함될 품목 라인 리스트.
        invoice_date: 명세서 발행일. None이면 오늘 날짜 사용.
        # supplier_name: 공급자명.
        # supplier_details: 공급자의 상세 정보 (사업자번호, 주소 등).

    Returns:
        생성된 Excel 파일의 경로. 실패 시 None.
    """
    wb = Workbook()
    ws = wb.active
    if ws is None:
        return None
    ws.title = "Invoice"

    # --- 스타일 정의 ---
    # 기본 폰트는 '맑은 고딕'으로 통일 시도
    font_name = "맑은 고딕"
    title_main_font = Font(name=font_name, size=20, bold=True) # 명세서 제목
    section_title_font = Font(name=font_name, size=11, bold=True) # (공급자), (공급받는자) 등
    header_font = Font(name=font_name, size=10, bold=True) # 테이블 헤더
    normal_font = Font(name=font_name, size=9)
    total_font = Font(name=font_name, size=10, bold=True) # 합계 폰트

    currency_format = "#,##0" # 원화 표시 없음, 숫자만
    
    thin_border_side = Side(style="thin", color="000000")
    medium_border_side = Side(style="medium", color="000000")
    
    table_cell_border = Border(
        left=thin_border_side, right=thin_border_side,
        top=thin_border_side, bottom=thin_border_side
    )
    total_row_border_top = Border(top=medium_border_side, bottom=medium_border_side, 
                                  left=thin_border_side, right=thin_border_side)
    total_row_label_border = Border(top=medium_border_side, bottom=medium_border_side, 
                                    left=medium_border_side, right=thin_border_side)
    total_row_value_border = Border(top=medium_border_side, bottom=medium_border_side, 
                                    left=thin_border_side, right=medium_border_side)


    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center", wrap_text=False)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # --- 문서 제목 (1행) ---
    current_dt = invoice_date if invoice_date else datetime.date.today()
    ws.merge_cells("A1:K1") # 11개 컬럼 사용 예정 (A-K)
    main_title_cell = ws["A1"]
    main_title_cell.value = f"{company.name} 납품 ({current_dt.strftime('%Y.%m.%d')})"
    main_title_cell.font = title_main_font
    main_title_cell.alignment = center_align
    ws.row_dimensions[1].height = 30

    # --- 공급자 / 공급받는자 정보 (3행부터) ---
    # 공급자 정보는 제거됨.
    # 공급받는자 정보만 표시.
    
    current_row = 3 # 시작 행 조정 (공급자 정보가 없으므로)
    
    # 공급받는자 정보
    ws.cell(row=current_row, column=1, value="공급받는자").font = section_title_font
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)

    ws.cell(row=current_row, column=3, value="상 호 명:").font = normal_font
    ws.cell(row=current_row, column=4, value=company.name).font = normal_font
    # 병합 범위를 늘려서 상호명이 길어도 잘 보이도록 조정 (예: K열까지)
    # 실제 컬럼 사용량(A-K)에 맞춰 end_column 조정 필요. 여기서는 예시로 F열까지.
    ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=6) # 기존 유지 또는 확장
    current_row += 2 # 테이블 전 공백 한 줄 (기존 current_row += 1 에서 변경)

    # --- 테이블 헤더 ---
    header_start_row = current_row
    table_headers = [
        "LOT", "모델명", "제품명", "규격", "납품수량", "단가", 
        "공급가액", "부가세", "보험수가", "치료재료코드", "UDI-DI"
    ]
    for col_idx, header_text in enumerate(table_headers, start=1):
        cell = ws.cell(row=header_start_row, column=col_idx, value=header_text)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = table_cell_border
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid") # 연한 회색 배경

    # --- 테이블 내용 ---
    current_row = header_start_row + 1
    total_qty = 0
    total_supply_amount = Decimal("0")
    total_vat = Decimal("0")

    for line in invoice_lines:
        col = 1
        ws.cell(row=current_row, column=col, value=line.lot).font = normal_font; col+=1
        ws.cell(row=current_row, column=col, value=line.model_name).font = normal_font; col+=1
        ws.cell(row=current_row, column=col, value=line.product_name).font = normal_font; col+=1
        ws.cell(row=current_row, column=col, value=line.spec).font = normal_font; col+=1
        
        qty_cell = ws.cell(row=current_row, column=col, value=line.qty)
        qty_cell.font = normal_font; qty_cell.alignment = right_align; col+=1
        
        unit_price_cell = ws.cell(row=current_row, column=col, value=line.unit_price)
        unit_price_cell.font = normal_font; unit_price_cell.number_format = currency_format; unit_price_cell.alignment = right_align; col+=1
        
        supply_amount_cell = ws.cell(row=current_row, column=col, value=line.supply_amount)
        supply_amount_cell.font = normal_font; supply_amount_cell.number_format = currency_format; supply_amount_cell.alignment = right_align; col+=1
        
        vat_cell = ws.cell(row=current_row, column=col, value=line.vat)
        vat_cell.font = normal_font; vat_cell.number_format = currency_format; vat_cell.alignment = right_align; col+=1
        
        insurance_price_val = line.insurance_price
        insurance_price_cell = ws.cell(row=current_row, column=col, value=insurance_price_val if insurance_price_val is not None else "")
        insurance_price_cell.font = normal_font; 
        if insurance_price_val is not None: insurance_price_cell.number_format = currency_format
        insurance_price_cell.alignment = right_align; col+=1

        ws.cell(row=current_row, column=col, value=line.treatment_code).font = normal_font; col+=1
        
        udi_di_to_write = str(line.udi_di) if line.udi_di is not None else ""
        udi_di_cell = ws.cell(row=current_row, column=col, value=udi_di_to_write)
        udi_di_cell.font = normal_font
        # Ensure it's treated as text by Excel if it looks like a number
        if udi_di_to_write.isdigit():
            udi_di_cell.number_format = '@' # Text format
        col+=1

        # 모든 데이터 셀에 테두리 적용
        for c_idx in range(1, len(table_headers) + 1):
            ws.cell(row=current_row, column=c_idx).border = table_cell_border
            # 기본 정렬 (텍스트는 왼쪽, 숫자는 오른쪽 - 위에서 개별 설정)
            if not ws.cell(row=current_row, column=c_idx).alignment.horizontal: # 이미 설정되지 않았다면
                 ws.cell(row=current_row, column=c_idx).alignment = left_align


        total_qty += line.qty
        total_supply_amount += line.supply_amount
        total_vat += line.vat
        current_row += 1

    # --- 합계 행 ---
    # "합계" 레이블 (LOT ~ 규격 컬럼 병합)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    total_label_cell = ws.cell(row=current_row, column=1, value="합계")
    total_label_cell.font = total_font
    total_label_cell.alignment = center_align
    total_label_cell.border = total_row_label_border # 왼쪽은 두꺼운 테두리

    # 총 납품수량
    total_qty_cell = ws.cell(row=current_row, column=5, value=total_qty)
    total_qty_cell.font = total_font; total_qty_cell.alignment = right_align
    total_qty_cell.border = total_row_border_top

    # 단가 컬럼은 비움
    ws.cell(row=current_row, column=6, value="").border = total_row_border_top


    # 총 공급가액
    total_supply_cell = ws.cell(row=current_row, column=7, value=total_supply_amount)
    total_supply_cell.font = total_font; total_supply_cell.number_format = currency_format; total_supply_cell.alignment = right_align
    total_supply_cell.border = total_row_border_top

    # 총 부가세
    total_vat_cell = ws.cell(row=current_row, column=8, value=total_vat)
    total_vat_cell.font = total_font; total_vat_cell.number_format = currency_format; total_vat_cell.alignment = right_align
    total_vat_cell.border = total_row_border_top
    
    # 나머지 합계행 컬럼 (보험수가, 치료재료코드, UDI-DI)은 비우고 테두리만 적용
    for c_idx in range(9, len(table_headers) + 1):
        cell = ws.cell(row=current_row, column=c_idx, value="")
        if c_idx == len(table_headers): # 마지막 셀
            cell.border = total_row_value_border # 오른쪽은 두꺼운 테두리
        else:
            cell.border = total_row_border_top


    # --- 열 너비 조정 ---
    column_widths = {
        'A': 15, 'B': 20, 'C': 30, 'D': 15, 'E': 8,  # LOT, 모델명, 제품명, 규격, 납품수량
        'F': 12, 'G': 12, 'H': 12, 'I': 12, # 단가, 공급가액, 부가세, 보험수가
        'J': 15, 'K': 25  # 치료재료코드, UDI-DI
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # --- 파일 저장 ---
    safe_company_name = "".join(c if c.isalnum() or c in " _-" else "" for c in company.name).strip()
    if not safe_company_name:
        safe_company_name = "invoice"
    
    filename_date_str = current_dt.strftime('%Y%m%d')
    base_filename = f"{filename_date_str}_{safe_company_name}_invoice"
    filename_ext = ".xlsx"
    
    # 파일명 중복 체크 및 넘버링
    counter = 1
    filename = f"{base_filename}{filename_ext}"
    filepath = os.path.join(os.getcwd(), filename)
    
    while os.path.exists(filepath):
        filename = f"{base_filename}_{counter}{filename_ext}"
        filepath = os.path.join(os.getcwd(), filename)
        counter += 1

    try:
        # 프로그램 실행 위치에 저장
        wb.save(filepath)
        print(f"거래명세서가 '{filepath}'에 저장되었습니다.")
        return filepath
    except Exception as e:
        print(f"Excel 파일 저장 중 오류 발생: {e}")
        return None

def open_file_explorer(filepath: str):
    """주어진 파일 경로를 파일 탐색기에서 엽니다. (파일을 선택한 상태로 폴더를 염)"""
    try:
        if platform.system() == "Windows":
            subprocess.run(['explorer', '/select,', os.path.normpath(filepath)], check=True)
        elif platform.system() == "Darwin": # macOS
            subprocess.run(['open', '-R', os.path.normpath(filepath)], check=True)
        else: # Linux and other Unix-like
            # xdg-open은 보통 디렉토리를 열지만, 파일 자체를 열 수도 있음.
            # 여기서는 파일이 있는 디렉토리를 여는 것이 목적.
            subprocess.run(['xdg-open', os.path.dirname(os.path.normpath(filepath))], check=True)
        print(f"파일 탐색기에서 '{filepath}' 위치를 열었습니다.")
    except FileNotFoundError: # explorer, open, xdg-open 등이 없을 경우
        print(f"오류: 파일 탐색기를 실행할 수 없습니다. 해당 명령어가 시스템에 설치되어 있는지 확인하세요.")
    except subprocess.CalledProcessError as e:
        print(f"파일 탐색기 실행 중 오류 발생: {e}")
    except Exception as e: # 기타 예외
        print(f"파일 탐색기 열기 중 예기치 않은 오류 발생: {e}")


if __name__ == '__main__':
    # 테스트용 데이터 (models.py의 구조에 맞게 수정)
    test_company = Company(id="comp-001", name="테스트 병원(주)", price_tier=PriceTier.A, contact="02-777-7777")
    
    item1_master = Item(
        lot="LOT2024A01", model_name="IMPLANT-X1", product_name="티타늄 스크류", spec="3.5mm x 10mm",
        treatment_code="C0012345", udi_di="880UDIITEM001X1",
        prices={PriceTier.A.value: Decimal("150000"), PriceTier.MEDICAL.value: Decimal("165000")}
    )
    item2_master = Item(
        lot="LOT2024B05", model_name="PLATE-Y2", product_name="정형외과 플레이트", spec="Type B, 5Hole",
        treatment_code="C0054321", udi_di="880UDIITEM002Y2",
        prices={PriceTier.A.value: Decimal("320000"), PriceTier.MEDICAL.value: Decimal("350000")}
    )
    item3_master = Item( # 보험수가가 없는 아이템
        lot="LOT2024C11", model_name="TOOL-Z3", product_name="수술용 가이드", spec="Standard",
        treatment_code="T0000001", udi_di="880UDIITEM003Z3",
        prices={PriceTier.A.value: Decimal("80000")}
    )

    # InvoiceLine 생성 시에는 해당 회사의 price_tier에 맞는 단가를 사용해야 함
    # 여기서는 test_company가 PriceTier.A 이므로, item_master.prices[PriceTier.A.value]를 unit_price로 사용
    
    invoice_lines_data = [
        InvoiceLine(item=item1_master, qty=2, unit_price=item1_master.get_price_for_tier(test_company.price_tier)),
        InvoiceLine(item=item2_master, qty=1, unit_price=item2_master.get_price_for_tier(test_company.price_tier)),
        InvoiceLine(item=item3_master, qty=5, unit_price=item3_master.get_price_for_tier(test_company.price_tier)),
    ]
    # unit_price가 None일 경우에 대한 방어 코드 (실제 앱에서는 UI 단에서 처리)
    for line in invoice_lines_data:
        if line.unit_price is None:
             line.unit_price = Decimal("0") # 테스트용으로 0 처리

    generated_filepath = create_invoice_excel(test_company, invoice_lines_data)

    if generated_filepath:
        print(f"\n테스트 Excel 파일 생성 완료: {generated_filepath}")
        # 생성된 파일 열기 (선택 사항)
        # open_file_explorer(generated_filepath) 
    else:
        print("\n테스트 Excel 파일 생성 실패.")
